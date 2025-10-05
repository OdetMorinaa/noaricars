from flask import Flask, render_template, redirect, url_for, request
from openpyxl import load_workbook
from datetime import datetime
import os
from apscheduler.schedulers.background import BackgroundScheduler

app = Flask(__name__)

# Fixed folder containing Excel files
FOLDER_PATH = r"C:\Users\anexr\Desktop\RENT A CAR NOARI"
os.makedirs(FOLDER_PATH, exist_ok=True)

# Global dictionary to store car status
car_status = {}

def parse_date(value):
    """Try to parse various date formats into a datetime object."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt)
            except:
                continue
    return None

def check_car_availability():
    """Scan Excel files and update car_status safely."""
    global car_status
    car_status.clear()

    for filename in os.listdir(FOLDER_PATH):
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            file_path = os.path.join(FOLDER_PATH, filename)
            car_info = {}
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active

                last_reserved = parse_date(ws["I3"].value)
                available_again = parse_date(ws["I4"].value)

                available = available_again and datetime.now() >= available_again

                car_info.update({
                    "last_reserved": last_reserved,
                    "available_again": available_again,
                    "available": available
                })
            except Exception as e:
                car_info["error"] = str(e)

            car_status[filename] = car_info

# APScheduler setup
scheduler = BackgroundScheduler()
scheduler.add_job(func=check_car_availability, trigger="interval", minutes=5)
scheduler.start()

@app.route("/")
def home():
    return render_template("index.html", car_status=car_status)

@app.route("/refresh")
def manual_refresh():
    check_car_availability()
    return redirect(url_for("home"))

@app.route("/edit/<filename>", methods=["GET", "POST"])
def edit_file(filename):
    file_path = os.path.join(FOLDER_PATH, filename)
    if not os.path.exists(file_path):
        return f"File {filename} not found", 404

    if request.method == "POST":
        # Get new dates from form
        last_reserved_str = request.form.get("last_reserved")
        available_again_str = request.form.get("available_again")

        # Parse the dates
        last_reserved = parse_date(last_reserved_str)
        available_again = parse_date(available_again_str)

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            ws["I3"].value = last_reserved.strftime("%d.%m.%Y") if last_reserved else None
            ws["I4"].value = available_again.strftime("%d.%m.%Y") if available_again else None

            wb.save(file_path)
            check_car_availability()
            return redirect(url_for("home"))
        except Exception as e:
            return f"Failed to save file: {e}", 500

    # GET request → show current dates
    car_info = car_status.get(filename, {})
    return render_template("edit.html", filename=filename, car_info=car_info)

if __name__ == "__main__":
    check_car_availability()
    app.run(host="0.0.0.0", port=5000, debug=True)
